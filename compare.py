import asyncio
import os
from telethon import TelegramClient
from telethon.tl.functions.channels import GetParticipantsRequest
from telethon.tl.functions.messages import GetCommonChatsRequest
from telethon.tl.types import ChannelParticipantsSearch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageDraw

# ========================================
#         ضع بياناتك هنا
# ========================================
api_id   = 39282826              # api_id من my.telegram.org
api_hash = '8b0e5652dcb650474c73477004b6d4ed'    # api_hash من my.telegram.org
phone    = '+201507783711'     # رقم تليفونك بالكود الدوبتاعك
# الجروب اللي عايز تتحقق إن الشخص فيه
TARGET_GROUP_ID = -1001679998238   # 4rd - FCAI 2022-2026

# الجروب اللي هتجيب أعضاءه وتفلترهم
SOURCE_GROUP_ID = -1001757252376
# ========================================

PHOTOS_FOLDER = 'photos'
THUMB_FOLDER  = 'photos_small'
THUMB_SIZE    = (40, 40)
ROW_HEIGHT    = 35
COL_A_WIDTH   = 7


async def get_all_members(client, group):
    all_participants = []
    offset = 0
    limit  = 200
    while True:
        participants = await client(GetParticipantsRequest(
            channel=group,
            filter=ChannelParticipantsSearch(''),
            offset=offset,
            limit=limit,
            hash=0
        ))
        if not participants.users:
            break
        all_participants.extend(participants.users)
        offset += len(participants.users)
    return all_participants


async def is_in_target_group_via_common(client, user, target_group_id):
    """تحقق عن طريق الجروبات المشتركة"""
    try:
        result = await client(GetCommonChatsRequest(
            user_id=user.id,
            max_id=0,
            limit=100
        ))
        for chat in result.chats:
            # مقارنة الـ ID بعد إزالة البادئة السالبة
            chat_id = int(f"-100{chat.id}") if chat.id > 0 else chat.id
            if chat_id == target_group_id or chat.id == abs(target_group_id):
                return True
        return False
    except Exception:
        return False


async def main():
    client = TelegramClient('session', api_id, api_hash)
    await client.start(phone=phone, password=lambda: "")
    print("✅ تم الاتصال بتليجرام بنجاح\n")

    # ── جيب الجروب المصدر ──
    source_group  = await client.get_entity(SOURCE_GROUP_ID)
    target_group  = await client.get_entity(TARGET_GROUP_ID)
    print(f"📌 الجروب المصدر : {source_group.title}")
    print(f"🎯 الجروب المستهدف: {target_group.title}\n")

    # ── جيب الأعضاء ──
    print("👥 جاري جلب الأعضاء...")
    members = await get_all_members(client, source_group)
    print(f"   إجمالي الأعضاء: {len(members)}\n")

    # ── فلترة عن طريق الجروبات المشتركة ──
    print("🔍 جاري التحقق من الجروبات المشتركة لكل عضو...")
    filtered = []

    for i, user in enumerate(members, 1):
        if user.bot:
            print(f"  [{i}/{len(members)}] ⚙️  {user.first_name or user.id} — بوت، تم تخطيه")
            continue

        in_group = await is_in_target_group_via_common(client, user, TARGET_GROUP_ID)

        if in_group:
            filtered.append(user)
            print(f"  [{i}/{len(members)}] ✅ {user.first_name or user.id} — مشترك في {target_group.title}")
        else:
            print(f"  [{i}/{len(members)}] ❌ {user.first_name or user.id} — مش مشترك")

        # استنى قليلاً عشان متتحجبش من تليجرام
        await asyncio.sleep(0.3)

    print(f"\n✅ الأعضاء المشتركين: {len(filtered)}")
    print(f"❌ تم حذف: {len(members) - len(filtered)} شخص\n")

    if not filtered:
        print("⚠️ مفيش أعضاء مشتركين. الكود هيوقف.")
        await client.disconnect()
        return

    # ── تحميل الصور ──
    os.makedirs(PHOTOS_FOLDER, exist_ok=True)
    os.makedirs(THUMB_FOLDER,  exist_ok=True)
    print("📸 جاري تحميل الصور...")

    photo_paths = {}
    for i, user in enumerate(filtered, 1):
        try:
            orig_path  = os.path.join(PHOTOS_FOLDER, f"{user.id}.jpg")
            thumb_path = os.path.join(THUMB_FOLDER,  f"{user.id}.png")

            if not os.path.exists(orig_path):
                downloaded = await client.download_profile_photo(user, file=orig_path)
            else:
                downloaded = orig_path

            if downloaded and os.path.exists(orig_path):
                img  = PILImage.open(orig_path).convert("RGBA")
                img  = img.resize(THUMB_SIZE, PILImage.LANCZOS)
                mask = PILImage.new("L", THUMB_SIZE, 0)
                draw = ImageDraw.Draw(mask)
                draw.ellipse((0, 0, THUMB_SIZE[0], THUMB_SIZE[1]), fill=255)
                img.putalpha(mask)
                img.save(thumb_path, "PNG")
                photo_paths[user.id] = thumb_path
                print(f"  [{i}/{len(filtered)}] ✅ {user.first_name or user.id}")
            else:
                print(f"  [{i}/{len(filtered)}] ⚪ {user.first_name or user.id} — مفيش صورة")
        except Exception as e:
            print(f"  [{i}/{len(filtered)}] ❌ {e}")

    # ── إنشاء Excel ──
    print("\n📊 جاري إنشاء ملف Excel...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = ['', '#', 'User ID', 'الاسم الأول', 'الاسم الأخير', 'الاسم الكامل', 'اليوزرنيم', 'رقم التليفون']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    col_widths = [COL_A_WIDTH, 5, 15, 15, 15, 25, 20, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for row, user in enumerate(filtered, 2):
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        username  = f"@{user.username}" if user.username else "—"
        phone_num = user.phone or "غير ظاهر"

        ws.row_dimensions[row].height = ROW_HEIGHT

        if user.id in photo_paths:
            try:
                img        = XLImage(photo_paths[user.id])
                img.width  = 32
                img.height = 32
                img.anchor = f"A{row}"
                ws.add_image(img)
            except Exception:
                ws.cell(row=row, column=1, value="—")
        else:
            ws.cell(row=row, column=1, value="—")

        ws.cell(row=row, column=2, value=row - 1)
        ws.cell(row=row, column=3, value=user.id)
        ws.cell(row=row, column=4, value=user.first_name or "")
        ws.cell(row=row, column=5, value=user.last_name  or "")
        ws.cell(row=row, column=6, value=full_name)
        ws.cell(row=row, column=7, value=username)
        ws.cell(row=row, column=8, value=phone_num)

        if row % 2 == 0:
            row_fill = PatternFill("solid", fgColor="EAF2FB")
            for c in range(2, 9):
                ws.cell(row=row, column=c).fill = row_fill

        for c in range(2, 9):
            ws.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')

    safe_name = source_group.title.replace("/", "-").replace("\\", "-")
    filename  = f"{safe_name}_FCAI_members.xlsx"
    wb.save(filename)

    print(f"\n🎉 تم! الملف: {filename}")
    print(f"👥 عدد الأعضاء في الملف: {len(filtered)}")
    await client.disconnect()

asyncio.run(main())